"""
BIRK FX Phase 2B Extended - Data Import Routes (FastAPI Version)
FastAPI endpoints for file uploads and manual exposure data entry
NOW WITH CRUD: Create, Read, Update, Delete
"""

from fastapi import APIRouter, HTTPException, UploadFile, File, Form, Query, Depends
from fastapi.responses import StreamingResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from pydantic import BaseModel, Field
from typing import Optional, List
from sqlalchemy.orm import Session
from datetime import datetime
import io
import csv
import os

from models import Exposure, Company, RiskLevel
from database import SessionLocal, get_live_fx_rate, calculate_risk_level

# ── Inline auth ───────────────────────────────────────────────────────────────
_security = HTTPBearer()

def _get_token_payload(credentials: HTTPAuthorizationCredentials = Depends(_security)) -> dict:
    from jose import JWTError, jwt
    SECRET_KEY = os.getenv("JWT_SECRET_KEY", "change-this-in-production-use-a-long-random-string")
    try:
        return jwt.decode(credentials.credentials, SECRET_KEY, algorithms=["HS256"])
    except JWTError:
        raise HTTPException(status_code=401, detail="Invalid or expired token")

router = APIRouter(prefix="/api/exposure-data", tags=["Exposure Data"])


# Pydantic Models
class ManualExposureRequest(BaseModel):
    company_id: int
    reference_number: str
    currency_pair: str
    amount: float = Field(..., gt=0)
    start_date: str
    end_date: str
    description: Optional[str] = None
    rate: Optional[float] = None
    # Budget & Risk Limits (Phase 2B)
    budget_rate: Optional[float] = None
    max_loss_limit: Optional[float] = None
    target_profit: Optional[float] = None
    hedge_ratio_policy: Optional[float] = Field(default=1.0, ge=0, le=1)
    instrument_type: Optional[str] = "Spot"
    amount_currency: Optional[str] = None  # Which currency the amount is in (defaults to from_currency)



# Database dependency
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


@router.post("/upload")
async def upload_file(
    file: UploadFile = File(...),
    company_id: int = Form(...),
    db: Session = Depends(get_db),
    payload: dict = Depends(_get_token_payload),
):
    """
    POST /api/exposure-data/upload

    Parses the Sumnohow FX Import Template (.xlsx only).
    Reads the sheet named "Exposures" and ignores all other sheets.

    Expected column headers (row 1):
      currency_pair | description | start_date | maturity_date |
      total_amount | budget_rate | instrument_type | base_currency

    Returns:
      {"imported": N, "skipped": N, "errors": ["Row 3: ...", ...]}
    """
    import openpyxl
    from datetime import date as _date

    try:
        # ── File type guard — xlsx only ───────────────────────────────────────
        filename = file.filename or ""
        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
        if ext != "xlsx":
            raise HTTPException(
                status_code=400,
                detail=f"Only .xlsx files are supported. Received: .{ext or 'unknown'}. "
                       f"Please use the Sumnohow Import Template."
            )

        # ── Size guard — 10 MB ────────────────────────────────────────────────
        contents = await file.read()
        if len(contents) > 10 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="File exceeds 10 MB limit.")

        # ── Multi-tenancy guard ───────────────────────────────────────────────
        company = db.query(Company).filter(Company.id == company_id).first()
        if not company:
            raise HTTPException(status_code=404, detail=f"Company {company_id} not found")
        role      = payload.get("role", "")
        token_cid = payload.get("company_id")
        if role not in ("superadmin", "admin", "company_admin") and int(token_cid or 0) != company_id:
            raise HTTPException(status_code=403, detail="Access denied")

        # ── Open workbook ─────────────────────────────────────────────────────
        try:
            wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Could not open .xlsx file: {e}")

        # Find the "Exposures" sheet (case-insensitive)
        sheet_name = next(
            (n for n in wb.sheetnames if n.strip().lower() == "exposures"),
            None,
        )
        if sheet_name is None:
            raise HTTPException(
                status_code=400,
                detail=f"Sheet named 'Exposures' not found. "
                       f"Available sheets: {', '.join(wb.sheetnames)}"
            )
        ws = wb[sheet_name]

        # ── Read and validate headers (row 1) ─────────────────────────────────
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise HTTPException(status_code=400, detail="'Exposures' sheet is empty.")

        raw_headers = rows[0]
        headers = [
            str(h).strip().lower() if h is not None else ""
            for h in raw_headers
        ]

        required_cols = {"currency_pair", "total_amount", "budget_rate", "instrument_type", "maturity_date", "start_date"}
        header_set    = set(headers)
        missing = required_cols - header_set
        if missing:
            raise HTTPException(
                status_code=400,
                detail=f"Missing required columns: {', '.join(sorted(missing))}. "
                       f"Please use the Sumnohow Import Template."
            )

        # Build column index lookup
        col = {name: headers.index(name) for name in headers if name}

        # ── Parse data rows ───────────────────────────────────────────────────
        from sqlalchemy import text as _text

        imported = 0
        skipped  = 0
        errors   = []
        today    = _date.today()

        VALID_INSTRUMENTS = {"forward", "spot", "option"}

        for row_idx, raw_row in enumerate(rows[1:], start=2):
            # Helper to read a cell by column name
            def cell(name):
                idx = col.get(name)
                if idx is None:
                    return ""
                val = raw_row[idx] if idx < len(raw_row) else None
                return str(val).strip() if val is not None else ""

            # Skip fully empty rows
            if not any(v for v in raw_row if v is not None):
                continue

            # ── currency_pair ─────────────────────────────────────────────────
            pair_raw = cell("currency_pair")
            if not pair_raw:
                skipped += 1
                continue   # silently skip blank rows

            # Accept both "EUR/USD" and "EURUSD"
            pair_norm = pair_raw.upper().replace("-", "")
            if "/" in pair_norm:
                parts = pair_norm.split("/")
                if len(parts) != 2 or len(parts[0]) != 3 or len(parts[1]) != 3:
                    errors.append(f"Row {row_idx}: invalid currency_pair '{pair_raw}' — must be XXX/XXX")
                    skipped += 1
                    continue
                from_currency, to_currency = parts[0], parts[1]
            else:
                if len(pair_norm) != 6:
                    errors.append(f"Row {row_idx}: invalid currency_pair '{pair_raw}' — must be XXX/XXX")
                    skipped += 1
                    continue
                from_currency = pair_norm[:3]
                to_currency   = pair_norm[3:]

            # ── total_amount ──────────────────────────────────────────────────
            amt_raw = cell("total_amount").replace(",", "")
            try:
                amount = float(amt_raw)
                if amount <= 0:
                    raise ValueError("must be > 0")
            except (ValueError, TypeError):
                errors.append(f"Row {row_idx}: invalid total_amount '{cell('total_amount')}' — must be a positive number")
                skipped += 1
                continue

            # ── budget_rate ───────────────────────────────────────────────────
            br_raw = cell("budget_rate")
            try:
                budget_rate = float(br_raw) if br_raw else None
                if budget_rate is not None and budget_rate <= 0:
                    raise ValueError("must be > 0")
            except (ValueError, TypeError):
                errors.append(f"Row {row_idx}: invalid budget_rate '{br_raw}' — must be a positive number")
                skipped += 1
                continue

            # ── start_date ────────────────────────────────────────────────────
            sd_raw = cell("start_date")
            if not sd_raw:
                errors.append(f"Row {row_idx}: start_date is required")
                skipped += 1
                continue
            try:
                raw_sd_val = raw_row[col["start_date"]] if col.get("start_date") is not None else None
                if hasattr(raw_sd_val, "date"):
                    start_date = raw_sd_val.date() if callable(raw_sd_val.date) else raw_sd_val
                elif hasattr(raw_sd_val, "year"):
                    start_date = raw_sd_val
                else:
                    start_date = datetime.strptime(sd_raw, "%Y-%m-%d").date()
            except (ValueError, TypeError, AttributeError):
                errors.append(f"Row {row_idx}: invalid start_date '{sd_raw}' — use YYYY-MM-DD")
                skipped += 1
                continue

            # ── maturity_date ─────────────────────────────────────────────────
            mat_raw = cell("maturity_date")
            try:
                # openpyxl may return a date object directly from a date cell
                raw_cell_val = raw_row[col["maturity_date"]] if col.get("maturity_date") is not None else None
                if hasattr(raw_cell_val, "date"):
                    maturity_date = raw_cell_val.date() if callable(raw_cell_val.date) else raw_cell_val
                elif hasattr(raw_cell_val, "year"):
                    # already a date object
                    maturity_date = raw_cell_val
                else:
                    maturity_date = datetime.strptime(mat_raw, "%Y-%m-%d").date()

                if maturity_date < today:
                    errors.append(f"Row {row_idx}: maturity_date '{mat_raw}' is in the past — must be >= today")
                    skipped += 1
                    continue
                if start_date >= maturity_date:
                    errors.append(f"Row {row_idx}: start_date must be before maturity_date")
                    skipped += 1
                    continue
            except (ValueError, TypeError, AttributeError):
                errors.append(f"Row {row_idx}: invalid maturity_date '{mat_raw}' — use YYYY-MM-DD")
                skipped += 1
                continue

            # ── instrument_type ───────────────────────────────────────────────
            inst_raw = cell("instrument_type")
            if inst_raw.lower() not in VALID_INSTRUMENTS:
                errors.append(
                    f"Row {row_idx}: invalid instrument_type '{inst_raw}' — "
                    f"must be one of: Forward, Spot, Option"
                )
                skipped += 1
                continue
            instrument_type = inst_raw.capitalize()

            # ── Optional fields ───────────────────────────────────────────────
            description   = cell("description")
            base_currency = cell("base_currency").upper() or from_currency
            amount_currency = base_currency if base_currency in (from_currency, to_currency) else from_currency

            # ── FX rate lookup ────────────────────────────────────────────────
            try:
                rate = get_live_fx_rate(from_currency, to_currency)
            except Exception:
                rate = 1.0
                errors.append(
                    f"Row {row_idx}: could not fetch live rate for {from_currency}/{to_currency} — "
                    f"defaulted to 1.0 (rate will need manual correction)"
                )

            period_days = (maturity_date - start_date).days
            usd_value   = amount * rate
            reference   = f"IMP-{row_idx:03d}"

            # ── Insert (savepoint per row so one failure never aborts others) ──
            # risk_level is intentionally omitted — column is nullable and the
            # PostgreSQL enum casing varies by how the DB was initialised.
            # The system recalculates risk_level elsewhere from amount + period.
            try:
                with db.begin_nested():   # savepoint — rolls back this row only on failure
                    db.execute(_text("""
                        INSERT INTO exposures
                          (company_id, from_currency, to_currency, amount, amount_currency,
                           start_date, end_date, initial_rate, current_rate, current_value_usd,
                           settlement_period, description, budget_rate,
                           reference, exposure_type, instrument_type, is_active, created_at, updated_at)
                        VALUES
                          (:cid, :from_ccy, :to_ccy, :amount, :amount_ccy,
                           :start_date, :end_date, :rate, :rate, :usd_value,
                           :period, :desc, :budget_rate,
                           :reference, 'payable', :instrument_type, true, NOW(), NOW())
                    """), {
                        "cid":             company_id,
                        "from_ccy":        from_currency,
                        "to_ccy":          to_currency,
                        "amount":          amount,
                        "amount_ccy":      amount_currency,
                        "start_date":      start_date,
                        "end_date":        maturity_date,
                        "rate":            rate,
                        "usd_value":       usd_value,
                        "period":          period_days,
                        "desc":            description,
                        "budget_rate":     budget_rate,
                        "reference":       reference,
                        "instrument_type": instrument_type,
                    })
                imported += 1
            except Exception as insert_err:
                errors.append(f"Row {row_idx}: database error — {insert_err}")
                skipped += 1
                continue

        db.commit()
        print(f"[upload] company_id={company_id} file='{filename}' imported={imported} skipped={skipped}")

        return {
            "success":  True,
            "filename": filename,
            "imported": imported,
            "skipped":  skipped,
            "errors":   errors,
            # Keep legacy field so existing frontend code doesn't break
            "row_count": imported,
            "summary": {
                "total_exposures": imported,
                "total_amount":    0,
                "unique_currencies": 0,
                "avg_period_days":   0,
            },
            "validation_warnings": errors,
        }

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")


@router.get("/template/{format}")
def download_template(format: str):
    """
    GET /api/exposure-data/template/csv   — dynamically generated CSV template
    GET /api/exposure-data/template/xlsx  — dynamically generated Excel template

    Dates are calculated relative to today so the sample rows are always
    future-dated and pass the maturity_date >= today validation rule.

    Column order matches the parser exactly:
      currency_pair | description | start_date | maturity_date |
      total_amount  | budget_rate | instrument_type | base_currency
    """
    from datetime import date, timedelta

    today = date.today()

    # Column headers — must match parser expected columns
    headers = [
        "currency_pair", "description", "start_date", "maturity_date",
        "total_amount", "budget_rate", "instrument_type", "base_currency",
    ]

    # Sample rows with future-relative dates so they always pass validation
    sample_rows = [
        ["GBP/USD", "Export receivables Q2",    today.isoformat(), (today + timedelta(45)).isoformat(),  3_000_000, 1.3200, "Forward", "GBP"],
        ["EUR/USD", "EU customer invoices",      today.isoformat(), (today + timedelta(60)).isoformat(),  5_000_000, 1.1600, "Forward", "EUR"],
        ["EUR/NOK", "Oslo office running costs", today.isoformat(), (today + timedelta(90)).isoformat(),  2_000_000, 11.200, "Forward", "EUR"],
        ["GBP/NOK", "Aberdeen supply contract",  today.isoformat(), (today + timedelta(75)).isoformat(),  8_000_000, 12.800, "Forward", "GBP"],
        ["CHF/USD", "Swiss supplier payments",   today.isoformat(), (today + timedelta(30)).isoformat(),  1_500_000, 1.2600, "Forward", "CHF"],
        ["USD/NOK", "US import costs H1",        today.isoformat(), (today + timedelta(120)).isoformat(), 4_000_000, 10.500, "Forward", "USD"],
    ]

    fmt = format.lower()

    if fmt == "csv":
        # Build CSV in-memory
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(headers)
        writer.writerows(sample_rows)
        return StreamingResponse(
            io.BytesIO(buf.getvalue().encode("utf-8")),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=Sumnohow_FX_Import_Template.csv"},
        )

    elif fmt in ("xlsx", "xls"):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()

        # ── Exposures sheet (the one the parser reads) ─────────────────────────
        ws = wb.active
        ws.title = "Exposures"

        # Header row — navy background, gold text, bold
        header_font  = Font(bold=True, color="C9A86C")
        header_fill  = PatternFill("solid", fgColor="1A2744")
        header_align = Alignment(horizontal="center")

        for col_idx, col_name in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align

        for row_idx, row in enumerate(sample_rows, start=2):
            for col_idx, val in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=val)

        # Auto-fit column widths
        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

        # ── Instructions sheet ─────────────────────────────────────────────────
        ws2 = wb.create_sheet("Instructions")
        instructions = [
            ["Sumnohow FX Import Template — Instructions"],
            [],
            ["IMPORTANT: Do not rename the 'Exposures' sheet. The parser reads that sheet by name."],
            [],
            ["Field",         "Description",                                                    "Required"],
            ["currency_pair", "Format: FROM/TO e.g. GBP/USD, EUR/NOK",                        "Yes"],
            ["description",   "Invoice ref, contract name, or counterparty",                   "Yes"],
            ["start_date",    "When exposure begins — trade or invoice date. Format: YYYY-MM-DD","Yes"],
            ["maturity_date", "When exposure settles — forward value date. YYYY-MM-DD. Must be after start_date and today or later", "Yes"],
            ["total_amount",  "Notional in the FROM currency. Numbers only, no commas",         "Yes"],
            ["budget_rate",   "Your internal planning rate e.g. 1.3200",                        "Yes"],
            ["instrument_type","Forward, Spot, or Option",                                      "Yes"],
            ["base_currency", "Your reporting currency e.g. EUR, GBP",                         "Yes"],
        ]
        for row_data in instructions:
            ws2.append(row_data)
        ws2.column_dimensions["A"].width = 18
        ws2.column_dimensions["B"].width = 70
        ws2.column_dimensions["C"].width = 10

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=Sumnohow_FX_Import_Template.xlsx"},
        )

    else:
        raise HTTPException(status_code=400, detail="Format must be csv or xlsx")

@router.post("/manual")
async def create_manual_exposure(
    request: ManualExposureRequest,
    db: Session = Depends(get_db)
):
    """
    POST /api/exposure-data/manual
    Create a single exposure manually
    """
    try:
        # Validate company exists
        company = db.query(Company).filter(Company.id == request.company_id).first()
        if not company:
            raise HTTPException(status_code=404, detail=f"Company {request.company_id} not found")
        
        # Parse currency pair (e.g., "EURUSD" -> from="EUR", to="USD")
        currency_pair = request.currency_pair.upper().replace("/", "").replace("-", "")
        
        if len(currency_pair) != 6:
            raise HTTPException(status_code=400, detail="Currency pair must be 6 characters (e.g., EURUSD)")
        
        from_currency = currency_pair[:3]
        to_currency = currency_pair[3:]
        
        # Calculate period days from dates
        start_date_obj = datetime.strptime(request.start_date, '%Y-%m-%d').date()
        end_date_obj = datetime.strptime(request.end_date, '%Y-%m-%d').date()
        period_days = (end_date_obj - start_date_obj).days
        
        if period_days <= 0:
            raise HTTPException(status_code=400, detail="End date must be after start date")
        
        # Get FX rate
        try:
            rate = request.rate if request.rate else get_live_fx_rate(from_currency, to_currency)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to get FX rate: {str(e)}")
        
        # Calculate USD value
        usd_value = request.amount * rate
        
        # Calculate risk level
        risk = calculate_risk_level(usd_value, period_days)
        
        # Create database record with Budget & Risk Limits
        db_exposure = Exposure(
            company_id=request.company_id,
            from_currency=from_currency,
            to_currency=to_currency,
            amount=request.amount,
            start_date=start_date_obj,
            end_date=end_date_obj,
            initial_rate=rate,
            current_rate=rate,
            current_value_usd=usd_value,
            settlement_period=period_days,
            risk_level=risk,
            description=request.description or '',
            budget_rate=request.budget_rate,
            max_loss_limit=request.max_loss_limit,
            target_profit=request.target_profit,
            hedge_ratio_policy=request.hedge_ratio_policy if request.hedge_ratio_policy else 1.0,
            instrument_type=request.instrument_type or 'Spot'
        )
        
        db.add(db_exposure)
        db.commit()
        db.refresh(db_exposure)

        # Save fields not on the ORM model via raw SQL (reference, exposure_type, amount_currency)
        from sqlalchemy import text as _text
        effective_amount_currency = (request.amount_currency or from_currency).upper()
        db.execute(_text("""
            UPDATE exposures
            SET reference        = :reference,
                exposure_type    = :exposure_type,
                amount_currency  = :amount_currency
            WHERE id = :id
        """), {
            "reference":       request.reference_number,
            "exposure_type":   getattr(request, 'exposure_type', 'payable') or 'payable',
            "amount_currency": effective_amount_currency,
            "id":              db_exposure.id,
        })
        db.commit()

        # Capture inception rate in the background (non-blocking)
        import asyncio as _asyncio
        from datetime import date as _date
        _asyncio.create_task(_capture_inception_rate(
            db_exposure.id, from_currency, to_currency, start_date_obj, rate
        ))

        return {
            'success': True,
            'exposure': {
                'id': db_exposure.id,
                'company_id': db_exposure.company_id,
                'reference_number': request.reference_number,
                'from_currency': db_exposure.from_currency,
                'to_currency': db_exposure.to_currency,
                'amount': db_exposure.amount,
                'start_date': db_exposure.start_date.isoformat() if db_exposure.start_date else None,
                'end_date': db_exposure.end_date.isoformat() if db_exposure.end_date else None,
                'initial_rate': db_exposure.initial_rate,
                'current_rate': db_exposure.current_rate,
                'current_value_usd': db_exposure.current_value_usd,
                'settlement_period': db_exposure.settlement_period,
                'risk_level': db_exposure.risk_level.value,
                'description': db_exposure.description,
                'budget_rate': db_exposure.budget_rate,
                'created_at': db_exposure.created_at.isoformat()
            },
            'message': f'Exposure {request.reference_number} created successfully'
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to create exposure: {str(e)}")



# NOTE: PUT /exposures/{exposure_id} is handled by the app-level endpoint in birk_api.py
# which supports budget_rate, instrument_type, exposure_type, from_currency, to_currency
# and enforces authentication. Do not add a duplicate router-level PUT here.


@router.delete("/exposures/{exposure_id}")
async def delete_exposure(
    exposure_id: int,
    db: Session = Depends(get_db)
):
    """
    DELETE /api/exposure-data/exposures/{exposure_id}
    Delete an exposure
    """
    try:
        # Fetch exposure
        exposure = db.query(Exposure).filter(Exposure.id == exposure_id).first()
        
        if not exposure:
            raise HTTPException(status_code=404, detail=f"Exposure {exposure_id} not found")
        
        # Store info for response
        exposure_info = {
            'id': exposure.id,
            'currency_pair': f"{exposure.from_currency}{exposure.to_currency}",
            'amount': exposure.amount
        }
        
        # Delete
        db.delete(exposure)
        db.commit()
        
        return {
            'success': True,
            'deleted_exposure': exposure_info,
            'message': f'Exposure {exposure_id} deleted successfully'
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to delete exposure: {str(e)}")


@router.get("/exposures/{company_id}")
async def get_company_exposures(
    company_id: int,
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    currency: Optional[str] = Query(None),
    status: str = Query(default="active"),
    db: Session = Depends(get_db)
):
    """
    GET /api/exposure-data/exposures/{company_id}
    Get all exposures for a company with optional filters
    """
    try:
        # Query exposures from database
        query = db.query(Exposure).filter(Exposure.company_id == company_id)
        
        # Apply filters if provided
        if currency:
            query = query.filter(
                (Exposure.from_currency == currency.upper()) | 
                (Exposure.to_currency == currency.upper())
            )
        
        exposures = query.all()
        
        # Convert to dict format
        exposure_list = []
        for exp in exposures:
            exposure_list.append({
                'id': exp.id,
                'company_id': exp.company_id,
                'from_currency': exp.from_currency,
                'to_currency': exp.to_currency,
                'currency_pair': f"{exp.from_currency}{exp.to_currency}",
                'amount': exp.amount,
                'start_date': exp.start_date.isoformat() if exp.start_date else None,
                'end_date': exp.end_date.isoformat() if exp.end_date else None,
                'initial_rate': exp.initial_rate,
                'current_rate': exp.current_rate,
                'current_value_usd': exp.current_value_usd,
                'settlement_period': exp.settlement_period,
                'period_days': exp.settlement_period,
                'risk_level': exp.risk_level.value if exp.risk_level else 'Unknown',
                'description': exp.description,
                'status': 'active',
                'created_at': exp.created_at.isoformat() if exp.created_at else None,
                'updated_at': exp.updated_at.isoformat() if exp.updated_at else None
            })
        
        return {
            'success': True,
            'company_id': company_id,
            'exposures': exposure_list,
            'count': len(exposure_list),
            'filters': {
                'start_date': start_date,
                'end_date': end_date,
                'currency': currency,
                'status': status
            }
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Server error: {str(e)}")


@router.get("/health")
async def health_check():
    """Health check endpoint"""
    return {
        'status': 'healthy',
        'service': 'Data Import',
        'version': '2.0.0',
        'features': ['CREATE', 'READ', 'UPDATE', 'DELETE']
    }

@router.get("/api/exposure-data/list")
async def list_manual_exposures(company_id: int, db: Session = Depends(get_db)):
    """Fetch all exposures for a company (matches manual entry endpoint pattern)"""
    from models import Exposure
    
    exposures = db.query(Exposure).filter(
        Exposure.company_id == company_id
    ).order_by(Exposure.created_at.desc()).limit(50).all()
    
    # Convert to dict for JSON serialization
    exposure_list = []
    for exp in exposures:
        exposure_list.append({
            "id": exp.id,
            "reference_number": exp.reference if hasattr(exp, 'reference') else exp.reference_number if hasattr(exp, 'reference_number') else None,
            "currency_pair": exp.currency_pair if hasattr(exp, 'currency_pair') else f"{exp.from_currency}{exp.to_currency}",
            "amount": float(exp.amount),
            "start_date": exp.start_date.isoformat() if hasattr(exp, 'start_date') and exp.start_date else None,
            "end_date": exp.end_date.isoformat() if hasattr(exp, 'end_date') and exp.end_date else None,
            "description": exp.description,
            "budget_rate": float(exp.budget_rate) if hasattr(exp, 'budget_rate') and exp.budget_rate else None,
            "hedge_ratio_policy": float(exp.hedge_ratio_policy) if hasattr(exp, 'hedge_ratio_policy') and exp.hedge_ratio_policy else 1.0,
            "created_at": exp.created_at.isoformat() if hasattr(exp, 'created_at') else None
        })
    
    return {"success": True, "exposures": exposure_list, "total": len(exposure_list)}


async def _capture_inception_rate(exposure_id: int, from_currency: str, to_currency: str, start_date, live_spot_rate: float):
    """
    Capture the inception rate for a newly created exposure.
    - Past date  → fetch ECB historical close rate
    - Today      → use the live spot already obtained at creation time
    - Future     → mark as 'ecb_scheduled' for the cron job to fill later
    """
    from datetime import date as _date
    from database import SessionLocal as _SessionLocal
    from sqlalchemy import text as _text
    from services.ecb_rates import get_cross_rate

    db = _SessionLocal()
    try:
        today = _date.today()

        if start_date > today:
            db.execute(_text("""
                UPDATE exposures SET inception_rate_source = 'ecb_scheduled' WHERE id = :id
            """), {"id": exposure_id})
            db.commit()
            return

        if start_date == today:
            rate   = live_spot_rate
            source = "live_spot"
        else:
            try:
                rate   = await get_cross_rate(from_currency, to_currency, start_date)
                source = "ecb_historical"
            except Exception as e:
                print(f"[inception] WARNING: ECB rate unavailable for {from_currency}/{to_currency} {start_date}: {e}")
                return

        db.execute(_text("""
            UPDATE exposures
            SET inception_rate = :rate,
                inception_rate_date = :dt,
                inception_rate_source = :src
            WHERE id = :id
        """), {"rate": rate, "dt": start_date, "src": source, "id": exposure_id})
        db.commit()
        print(f"[inception] captured {from_currency}/{to_currency} {start_date}: {rate:.6f} ({source})")
    except Exception as e:
        print(f"[inception] ERROR capturing for exposure {exposure_id}: {e}")
    finally:
        db.close()